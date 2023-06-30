from tkinter import *
from MainWindow.Controls.Select_Path import Path_Selection
from MainWindow.Controls.Edit_Buttons import EditButtons
from MainWindow.Controls.Final_Buttons import *
import openpyxl 
from PIL import Image, ImageTk, ImageDraw , ImageFont
from tkinter import font
import shutil
import os




class Control_Pane(Frame):
    

    def __init__(self, parent = None ):
        Frame.__init__(self, parent)
        super().__init__(parent)
        self.config(bg= '#2d2d32')


        Edit_button = EditButtons(self)
        path_selection = Path_Selection(self)
        final_but = Final_Buttons(self)

        self.response = Label(self, text= ' ', font= (font.Font(family= 'arial', size= 20, weight= 'bold')), fg= 'white', bg= '#2d2d32', width= 15, height= 5)

        img = Image.open('logo1.png')
        self.logoimg = ImageTk.PhotoImage(img)

        logo = Label(self, image= self.logoimg, border= 0)
        
        

        
        
        x_pad = 3
        y_pad = 4


        path_selection.grid(row= 0, column= 0, padx= x_pad, pady= y_pad)
        Edit_button.grid(row= 1, column= 0, padx= x_pad)
        final_but.grid(row=2, column= 0, padx= x_pad, pady= y_pad)

        self.response.grid(row= 3, column= 0)
        logo.grid(row= 4, column= 0)

        self.Edit_button = Edit_button
        self.path_selection = path_selection
        self.final_but = final_but
        self.coordinate_list = []
        self.Entry_list = []
        self.primary_list = []
        self.colour_list = []
        self.font_size_list = []
        self.font_family_list = []

        self.count = 0
    
    def set_certificate(self, certificate = None):
        self.certificate = certificate
        self.Temp_Folder = certificate
        
        self.master.preview_frame.RenderImage(certificate)

        print(certificate)
    
    



    def Get_Text_List(self, row , column):
        print('name list generated')
        
        
        path = self.Excel
        text_list = []
        

        work_book = openpyxl.load_workbook(path)
        sheet = work_book.active


        for i in range(row, sheet.max_row  , 1):

            text = sheet.cell(column = column,  row = i  ).value
            if(text == None):
                break
            
            text_list.append(text)
            
        
 
        if(self.count == 0):
            self.primary_list = text_list
            self.count = self.count + 1
        else:
            self.count = self.count + 1
        
        return text_list
    
    def make_position_list(self):
        
        x1, y1, x2, y2 = self.coordinate

        abcissa = x1 * self.ratio
        ordinate = y1 * self.ratio
        self.coordinate_list.append((abcissa, ordinate))

    
    def make_Entry_list(self, text_list):
        self.Entry_list.append(text_list)
    
    def make_font_style_list(self, font_style):
        self.font_family_list.append(font_style)
    
    def make_font_size_list(self, font_size):
        self.font_size_list.append(font_size)
    
    def make_font_color_list(self, color):
        self.colour_list.append(color)

    
    def set_primary_list(self, text_list):
        self.primary_list = text_list
        print(self.primary_list)

    

    def set_Excel(self, Excel):
        self.Excel = Excel
        
        
        
    
    def set_output(self, output):
        self.output = output
        
        

    def set_font(self, __font__):
        self._font = __font__
        
    
    def set_coordinate(self, coordinate, small_width , small_height ):

        self.coordinate = coordinate
        self.small_width = small_width
        self.small_height = small_height
        


        print(coordinate)
    

    def paste_png(self, file_path):

        print(self.Temp_Folder)
        image1 = Image.open(self.Temp_Folder)
        image2 = Image.open(file_path)

        
        width , height = image1.size
        self.ratio = width/self.small_width

        x1 = int(self.coordinate[0]*self.ratio)
        y1 = int(self.coordinate[1]*self.ratio)
        x2 = int(self.coordinate[2]*self.ratio)
        y2 = int(self.coordinate[3]*self.ratio)

        box_width = x2 -x1
        box_height = y2 -y1

        img2_width , img2_height = image2.size

        max_size = max(img2_width, img2_height)

        img2_ratio = img2_width/ img2_height

        new_img_max = int(min(box_width, box_height)*img2_ratio)

        img2_width = int(img2_width * new_img_max /max_size)
        img2_height = int(img2_height * new_img_max /max_size)

        print(box_width, box_height)
        print(img2_width, img2_height)

        image2 = image2.resize((img2_width, img2_height))

        image1 = image1.convert('RGBA')
        image2 = image2.convert('RGBA')

        image1.paste(image2,(x1, y1), image2)
        image1 = image1.convert('RGB')



   
   
        

        image1.save('Background.jpg')


        self.master.preview_frame.RenderImage('Background.jpg')
    


    def Generate_1_image(self, text = None):
        print(self.Temp_Folder)
        print(self.Temp_Folder)
        print(self.Excel)
        print(self.output)
        print(self._font)
        self.Font_folder = 'Font'
        
        sys_family = self._font[0]
        sys_size = int(self._font[1])
        color = self._font[2]
        img = Image.open(self.Temp_Folder)


        width , height = img.size

        self.ratio = width/self.small_width
        

        x1, y1, x2, y2 = self.coordinate

        abcissa = x1 * self.ratio
        ordinate = y1 * self.ratio
        self.abcissa = abcissa

        self.ordinate = ordinate
        print(self.ratio)
        print(abcissa, ordinate)

        __font = ImageFont.truetype(font= (self.Font_folder + '/'+ sys_family) , size= int(sys_size * self.ratio))
        
        print("One Image Generated")
        
        if(text == None):
            text = "Student Name"

        draw = ImageDraw.Draw(img)
        draw.text((abcissa, self.ordinate),text= text,fill= self._font[2],  font=__font, align= 'center'  )

        

        

        img.save('Background.jpg')


        self.master.preview_frame.RenderImage('Background.jpg')

    def Generate_All(self):
        pass
        print(self.coordinate_list)
        print(self.Entry_list)
        print(self.primary_list)
        print(self.font_family_list)
        print(self.font_size_list)
        print(self.colour_list)

        if(os.listdir('PreOutput/') == []):
            shutil.copy(self.certificate, 'PreOutput/')
            temp_image = os.listdir('PreOutput/')[0]
            print(temp_image)
            os.rename('PreOutput/' + temp_image, 'PreOutput/Background.jpg')
            print(os.listdir('PreOutput/'))

            
        for j in range(0,len(self.Entry_list[0]), 1):
            img = Image.open('PreOutput/Background.jpg')
            I1 = ImageDraw.Draw(img)
            for i  in range(0, len(self.coordinate_list), 1):

                print(self.Entry_list[i][j], self.colour_list[i], self.font_family_list[i], self.font_size_list[i], self.coordinate_list[i])

                __font = ImageFont.truetype(font= (self.Font_folder + '/'+ self.font_family_list[i]) , size= int( int(self.font_size_list[i]) * self.ratio))

                I1.text(self.coordinate_list[i], text= self.Entry_list[i][j] ,fill= self.colour_list[i], font= __font)

            img.save(self.output+ '/' + self.primary_list[j] + '.jpg')

        os.remove('PreOutput/Background.jpg')


        



    
    def start(self):
        self.mainloop()








