from tkinter import *
from Select_Path import Path_Selection
from Font_and_Color import *
from Final_Buttons import *
import openpyxl 
from PIL import Image, ImageTk, ImageDraw , ImageFont
from tkinter import font




class Control_Pane(Frame):

    def __init__(self, parent = None ):
        Frame.__init__(self, parent)
        super().__init__(parent)



        Font_and_Color = Font_And_Color(self)
        path_selection = Path_Selection(self)
        final_but = Final_Buttons(self)


        path_selection.grid(row= 0, column= 0)
        Font_and_Color.grid(row= 1, column= 0)
        final_but.grid(row=2, column= 0)

        self.Font_and_color = Font_and_Color
        self.path_selection = path_selection
        self.final_but = final_but
    
    def set_certificate(self, certificate = None):
        self.certificate = certificate
        
        self.master.preview_frame.RenderImage(certificate)

        print(certificate)
    def Get_Name_List(self):
        sheet_column = 1
        path = self.Excel
        name_list = []

        work_book = openpyxl.load_workbook(path)
        sheet = work_book.active


        for i in range(1, sheet.max_row + 1 , 1):
            name = sheet.cell(column = sheet_column,  row = i  ).value
            name_list.append(name)
        
        return name_list

    def set_Excel(self, Excel):
        self.Excel = Excel
        self.NameList = self.Get_Name_List()
        print(self.NameList)
        
    
    def set_output(self, output):
        self.output = output
        
        

    def set_font(self, __font__):
        self._font = __font__
        
    
    def set_coordinate(self, coordinate, small_width , small_height ):
        self.coordinate = coordinate
        self.small_width = small_width
        self.small_height = small_height
        print(coordinate)

    def Generate_1_image(self):
        print(self.certificate)
        print(self.Excel)
        print(self.output)
        print(self._font)
        
        sys_family = self._font[1]
        sys_size = int(self._font[2])
        font_weight = self._font[3]
        img = Image.open(self.certificate)

        width , height = img.size

        ratio = int(width/self.small_width )

        x1, y1, x2, y2 = self.coordinate

        abcissa = x1 * ratio
        ordinate = y2 * ratio
        self.abcissa = abcissa
        self.ordinate = ordinate

        __font = ImageFont.truetype(font= 'C:/Users/akash/Desktop/Certificate Naming Project/Font/Lato-Regular.ttf' , size= sys_size*5)
        
        print("One Image Generated")

        
        draw = ImageDraw.Draw(img)
        draw.text((abcissa, ordinate), "Student Name",fill= self._font[0],  font=__font,  )
        img.save('Test.jpg')
        self.master.preview_frame.RenderImage('Test.jpg')

    def Generate_All(self):

        sys_size = int(self._font[2])
        __font = ImageFont.truetype(font= 'C:/Users/akash/Desktop/Certificate Naming Project/Font/Lato-Regular.ttf' , size= sys_size*5)
        
        for name in self.NameList:
            img = Image.open(self.certificate)
            I1 = ImageDraw.Draw(img)
            I1.text((self.abcissa, self.ordinate), text= name ,fill= self._font[0], font= __font)
            img.save(self.output+ '/' + name + '.jpg')
        
        print('Done')



    
    def start(self):
        self.mainloop()








