from Modules import getPath 

import openpyxl as excel


#Getting Address of Every File Needed
def get_paths():
    Certificate = getPath.AskPath()
    Font = getPath.AskPath()
    Output = getPath.AskFolder()

    return(Certificate, Font , Output)



def Get_Name_List( ):
    sheet_column = int(input("Enter column number : "))
    path = getPath.AskPath()
    name_list = []

    work_book = excel.load_workbook(path)
    sheet = work_book.active


    for i in range(1, sheet.max_row + 1 , 1):
        name = sheet.cell(column = sheet_column,  row = i  ).value
        name_list.append(name)
    
    return name_list


def get_color():
    red = int(input("Enter Red code of color: ")) % 255
    green = int(input("Enter Green of color: ")) % 255
    blue = int(input("Enter Green of color: ")) % 255
    return (red, green, blue)



# left_x , left_y, right_x,right_y =GetCordinates.Position(Certficate_Path)
# print(left_x , left_y, right_x , right_y)
